import os
import sys
import time
import multiprocessing as mp

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import openpyxl
from openpyxl.utils import get_column_letter

from src.config import Config
from src.loader import TpmLoader
from src.strategies.geometric import GeometricSIA
from src.strategies.q_nodes import QNodes


TIMEOUT_SEC = 120
EXCEL_PATH = os.path.join(
    os.path.dirname(__file__), "..", "data", "DatosPruebas2026.xlsx"
)

ESTRATEGIAS = {"QNodes": QNodes, "Geometric": GeometricSIA}


def _deshabilitar_profiler():
    import src.middlewares.profile as pm

    pm.profiler_manager.enabled = False


def _string_a_mask(sistema, subconjunto):
    return "".join("1" if ch in subconjunto else "0" for ch in sistema)


def _ejecutar_estrategia(cola, N, pagina, estado_inicial, condiciones, alcance, mecanismo, nombre_estrategia):
    try:
        _deshabilitar_profiler()
        tpm = TpmLoader.cargar(N, pagina)
        config = Config(pagina_muestra=pagina)
        Cls = ESTRATEGIAS[nombre_estrategia]
        analizador = Cls(tpm, config)
        inicio = time.perf_counter()
        solucion = analizador.aplicar_estrategia(
            estado_inicial, condiciones, alcance, mecanismo
        )
        elapsed = time.perf_counter() - inicio
        cola.put(("OK", elapsed, float(solucion.perdida), solucion.particion))
    except Exception as e:
        import traceback

        tb = traceback.format_exc()
        cola.put(("ERROR", 0.0, 0.0, f"{type(e).__name__}: {e}"))


def _medir(N, pagina, estado_inicial, condiciones, alcance, mecanismo, nombre_estrategia):
    cola = mp.Queue()
    args = (cola, N, pagina, estado_inicial, condiciones, alcance, mecanismo, nombre_estrategia)
    proceso = mp.Process(target=_ejecutar_estrategia, args=args)
    proceso.start()
    proceso.join(timeout=TIMEOUT_SEC)

    if proceso.is_alive():
        proceso.terminate()
        proceso.join()
        return "TIMEOUT", TIMEOUT_SEC, 0.0, ""

    try:
        estado, valor, extra, particion = cola.get_nowait()
    except Exception:
        return "ERROR", 0.0, 0.0, ""

    return estado, valor, extra, particion


def procesar_sheet(ws, N, pagina, estado_inicial_str):
    sistema = ws["B2"].value
    if not sistema:
        print(f"  -> No se encontro Sistema en B2, saltando")
        return

    N_expected = len(sistema)
    if N_expected != N:
        print(f"  -> Sistema tiene {N_expected} nodos, esperaba {N}, saltando")
        return

    condiciones = "1" * N

    fila = 6
    total_ok = 0
    total_err = 0

    while True:
        num_prueba = ws.cell(row=fila, column=1).value
        if num_prueba is None:
            break

        alcance_str = ws.cell(row=fila, column=2).value
        mecanismo_str = ws.cell(row=fila, column=3).value

        if not alcance_str or not mecanismo_str:
            fila += 1
            continue

        alcance = _string_a_mask(sistema, alcance_str)
        mecanismo = _string_a_mask(sistema, mecanismo_str)

        print(f"  Prueba {num_prueba}: alcance={alcance_str} mech={mecanismo_str}", end="")

        for estrategia, col_start in [("QNodes", 4), ("Geometric", 7)]:
            estado, t, perdida, particion = _medir(
                N, pagina, estado_inicial_str, condiciones, alcance, mecanismo, estrategia
            )

            if estado == "OK":
                ws.cell(row=fila, column=col_start, value=particion)
                ws.cell(row=fila, column=col_start + 1, value=float(perdida))
                ws.cell(row=fila, column=col_start + 2, value=round(t, 6))
                total_ok += 1
            else:
                ws.cell(row=fila, column=col_start, value=f"ERROR: {particion}")
                ws.cell(row=fila, column=col_start + 1, value=0.0)
                ws.cell(row=fila, column=col_start + 2, value=0.0)
                total_err += 1

            print(f" | {estrategia}:{t:.3f}s", end="")

        print()
        fila += 1

    print(f"  -> {total_ok} OK, {total_err} ERROR")


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH)

    configs = [
        ("10A-Elementos", 10, "A", "1000000000"),
        ("15B-Elementos", 15, "B", "100000000000000"),
    ]

    for sheet_name, N, pagina, estado_inicial_str in configs:
        if sheet_name not in wb.sheetnames:
            print(f"Hoja '{sheet_name}' no encontrada, saltando")
            continue

        ws = wb[sheet_name]
        print(f"\n=== {sheet_name} ===")
        procesar_sheet(ws, N, pagina, estado_inicial_str)

    out_path = EXCEL_PATH
    wb.save(out_path)
    print(f"\nExcel guardado en: {out_path}")


if __name__ == "__main__":
    mp.freeze_support()
    main()
