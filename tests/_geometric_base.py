import os
import sys
import time

os.environ["PYPHI_WELCOME_OFF"] = "yes"

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from src.config import Config
from src.loader import TpmLoader
from src.strategies.k_geometric import KGeometric


def _deshabilitar_profiler():
    import src.middlewares.profile as pm

    pm.profiler_manager.enabled = False


def letters_to_mask(sistema: str, subset: str) -> str:
    return "".join("1" if ch in subset else "0" for ch in sistema)


def run_scenario(N, page, estado_inicial, alcance_mask, mecanismo_mask, condiciones_mask, k):
    _deshabilitar_profiler()
    tpm = TpmLoader.cargar(N, page)
    config = Config(pagina_muestra=page)
    analizador = KGeometric(tpm, config, k=k)

    inicio = time.perf_counter()
    soluciones = analizador.aplicar_estrategia(
        estado_inicial, condiciones_mask, alcance_mask, mecanismo_mask,
    )
    elapsed = time.perf_counter() - inicio

    mejor = soluciones[0]
    return mejor.particion, float(mejor.perdida), elapsed


K_COLUMNS = {
    2: (7, 8, 9),
    3: (13, 14, 15),
    4: (19, 20, 21),
    5: (25, 26, 27),
}


def process_sheet(excel_path, sheet_name, N, page):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name]

    sistema = ws["B2"].value
    estado_inicial = str(ws["B1"].value)
    condiciones_mask = "1" * N

    total_scenarios = 0
    for row in range(6, 56):
        alcance_letters = ws[f"B{row}"].value
        mecanismo_letters = ws[f"C{row}"].value

        if alcance_letters is None or mecanismo_letters is None:
            break

        total_scenarios += 1
        alcance_mask = letters_to_mask(sistema, alcance_letters)
        mecanismo_mask = letters_to_mask(sistema, mecanismo_letters)

        print(f"\n[{total_scenarios:2d}] Row {row}:")
        print(f"     alcance={alcance_letters} ({len(alcance_letters)}ch)")
        print(f"     mecanismo={mecanismo_letters} ({len(mecanismo_letters)}ch)")

        for k in (2, 3, 4, 5):
            part_col, loss_col, time_col = K_COLUMNS[k]

            try:
                particion_str, perdida, elapsed = run_scenario(
                    N, page,
                    estado_inicial, alcance_mask, mecanismo_mask,
                    condiciones_mask, k,
                )

                ws.cell(row=row, column=part_col, value=particion_str)
                ws.cell(row=row, column=loss_col, value=perdida)
                ws.cell(row=row, column=time_col, value=round(elapsed, 6))

                print(f"     k={k}: phi={perdida:.6f}  t={elapsed:.4f}s")

            except Exception as e:
                print(f"     k={k}: ERROR - {e}")
                ws.cell(row=row, column=part_col, value=f"ERROR: {e}")
                ws.cell(row=row, column=loss_col, value=None)
                ws.cell(row=row, column=time_col, value=None)

    wb.save(excel_path)
    print(f"\nProcesados {total_scenarios} escenarios. Guardado: {excel_path}")
