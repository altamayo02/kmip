import concurrent.futures
import os
import sys
import time

os.environ["PYPHI_WELCOME_OFF"] = "yes"

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from src.config import Config
from src.loader import TpmLoader
from src.strategies.phi import Phi


def _deshabilitar_profiler():
    import src.middlewares.profile as pm

    pm.profiler_manager.enabled = False


def letters_to_mask(sistema: str, subset: str) -> str:
    return "".join("1" if ch in subset else "0" for ch in sistema)


K_COLUMNS = {
    2: (4, 5, 6),
    3: (10, 11, 12),
    4: (16, 17, 18),
    5: (22, 23, 24),
}


def _workers_for_n(N):
    if N <= 10:
        return 8
    elif N <= 15:
        return 6
    elif N <= 20:
        return 4
    elif N <= 22:
        return 3
    return 2


def _process_row(args):
    N, page, estado_inicial, condiciones_mask, alcance_mask, mecanismo_mask, row_idx = args

    _deshabilitar_profiler()
    tpm = TpmLoader.cargar(N, page)
    config = Config(pagina_muestra=page)

    results = []
    for k in (2, 3, 4, 5):
        try:
            analizador = Phi(tpm, config, k=k)
            inicio = time.perf_counter()
            soluciones = analizador.aplicar_estrategia(
                estado_inicial, condiciones_mask, alcance_mask, mecanismo_mask,
            )
            elapsed = time.perf_counter() - inicio
            mejor = soluciones[0]
            results.append({
                'k': k,
                'particion': mejor.particion,
                'perdida': float(mejor.perdida),
                'tiempo': elapsed,
            })
        except Exception as e:
            results.append({
                'k': k,
                'error': str(e),
            })
    return row_idx, results


def _preparar_hoja(wb, sheet_name, template_sheet_name):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    print(f"Creando hoja '{sheet_name}' desde plantilla '{template_sheet_name}'...")
    template = wb[template_sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    for src_row in (1, 2):
        for col in (1, 2):
            ws.cell(row=src_row, column=col, value=template.cell(row=src_row, column=col).value)

    for row in range(6, 56):
        b_val = template.cell(row=row, column=2).value
        c_val = template.cell(row=row, column=3).value
        if b_val is None or c_val is None:
            break
        ws.cell(row=row, column=2, value=b_val)
        ws.cell(row=row, column=3, value=c_val)

    return ws


def process_sheet(excel_path, sheet_name, N, page, template_sheet_name=None):
    wb = openpyxl.load_workbook(excel_path)
    ws = _preparar_hoja(wb, sheet_name, template_sheet_name)

    sistema = ws["B2"].value
    estado_inicial = str(ws["B1"].value)
    condiciones_mask = "1" * N

    tasks = []
    for row in range(6, 56):
        alcance_letters = ws[f"B{row}"].value
        mecanismo_letters = ws[f"C{row}"].value
        if alcance_letters is None or mecanismo_letters is None:
            break
        alcance_mask = letters_to_mask(sistema, alcance_letters)
        mecanismo_mask = letters_to_mask(sistema, mecanismo_letters)
        tasks.append((N, page, estado_inicial, condiciones_mask, alcance_mask, mecanismo_mask, row))

    total = len(tasks)
    w = _workers_for_n(N)
    print(f"Procesando {total} escenarios con {w} workers (N={N})...")

    with concurrent.futures.ProcessPoolExecutor(max_workers=w) as executor:
        future_map = {executor.submit(_process_row, t): t[-1] for t in tasks}
        completed = 0

        for future in concurrent.futures.as_completed(future_map):
            row_idx, results = future.result()
            completed += 1

            alcance_letters = ws[f"B{row_idx}"].value
            mecanismo_letters = ws[f"C{row_idx}"].value
            print(f"\n[{completed:2d}] Row {row_idx}:")
            print(f"     alcance={alcance_letters} ({len(alcance_letters)}ch)")
            print(f"     mecanismo={mecanismo_letters} ({len(mecanismo_letters)}ch)")

            for res in results:
                k = res['k']
                part_col, loss_col, time_col = K_COLUMNS[k]

                if 'error' in res:
                    print(f"     k={k}: ERROR - {res['error']}")
                    ws.cell(row=row_idx, column=part_col, value=f"ERROR: {res['error']}")
                    ws.cell(row=row_idx, column=loss_col, value=None)
                    ws.cell(row=row_idx, column=time_col, value=None)
                else:
                    ws.cell(row=row_idx, column=part_col, value=res['particion'])
                    ws.cell(row=row_idx, column=loss_col, value=res['perdida'])
                    ws.cell(row=row_idx, column=time_col, value=round(res['tiempo'], 6))
                    print(f"     k={k}: phi={res['perdida']:.6f}  t={res['tiempo']:.4f}s")

            if completed % 10 == 0:
                wb.save(excel_path)
                print(f"  (guardado parcial: {completed}/{total})")

    wb.save(excel_path)
    print(f"\nProcesados {total} escenarios. Guardado: {excel_path}")
