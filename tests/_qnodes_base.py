import os
import sys
import time

os.environ["PYPHI_WELCOME_OFF"] = "yes"
os.environ.setdefault("PYTHONIOENCODING", "utf-8")

# Suppress CuPy CUDA_PATH warning by pre-setting from pip-installed nvidia packages
if not os.environ.get("CUDA_PATH"):
    for _sp in sys.path:
        _candidate = os.path.join(_sp, "nvidia", "cuda_runtime")
        if os.path.isfile(os.path.join(_candidate, "include", "cuda_runtime.h")):
            os.environ["CUDA_PATH"] = _candidate
            break

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from src.config import Config
from src.loader import TpmLoader
from src.strategies.k_q_nodes import KQNodes
from src.functions.gpu_backend import HAS_GPU


def _deshabilitar_profiler():
    import src.middlewares.profile as pm
    pm.profiler_manager.enabled = False


def letters_to_mask(sistema: str, subset: str) -> str:
    return "".join("1" if ch in subset else "0" for ch in sistema)


K_COLUMNS = {
    2: (4, 5, 6),
    3: (10, 11, 12),
    4: (16, 17, 18),
    5: (22, 23, 24),
}


_TPM_CACHE: dict = {}
_subsistema_cache: dict = {}


def _procesar_escenario(N, page, estado_inicial, condiciones_mask, alcance_mask, mecanismo_mask, use_gpu):
    cache_key = (N, page)
    tpm = _TPM_CACHE.get(cache_key)
    if tpm is None:
        tpm = TpmLoader.cargar(N, page)
        _TPM_CACHE[cache_key] = tpm

    config = Config(pagina_muestra=page)

    # Prepare subsystem once (shared across all k values)
    preparador = KQNodes(tpm, config, k=2, use_gpu=use_gpu)
    preparador.sia_preparar_subsistema(estado_inicial, condiciones_mask, alcance_mask, mecanismo_mask)
    shared = (preparador.sia_subsistema, preparador.sia_dists_marginales)

    results = []
    for k in (2, 3, 4, 5):
        try:
            analizador = KQNodes(tpm, config, k=k, use_gpu=use_gpu)
            analizador.sia_subsistema = shared[0]
            analizador.sia_dists_marginales = shared[1]
            analizador.sia_tiempo_inicio = time.time()

            inicio = time.perf_counter()
            soluciones = analizador.aplicar_estrategia(
                estado_inicial, condiciones_mask, alcance_mask, mecanismo_mask,
                _skip_prep=True,
            )
            elapsed = time.perf_counter() - inicio
            mejor = soluciones[0]
            results.append({
                'k': k,
                'particion': mejor.particion,
                'perdida': float(mejor.perdida),
                'tiempo': elapsed,
            })
        except Exception as e:
            results.append({
                'k': k,
                'error': str(e),
            })
    return results


def process_sheet(excel_path, sheet_name, N, page):
    _deshabilitar_profiler()
    use_gpu = HAS_GPU
    modo = "GPU (CUDA)" if use_gpu else "CPU"
    print(f"Modo: {modo}")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name]

    sistema = ws["B2"].value
    estado_inicial = str(ws["B1"].value)
    condiciones_mask = "1" * N

    escenarios = []
    for row in range(6, 56):
        alcance_letters = ws[f"B{row}"].value
        mecanismo_letters = ws[f"C{row}"].value
        if alcance_letters is None or mecanismo_letters is None:
            break
        alcance_mask = letters_to_mask(sistema, alcance_letters)
        mecanismo_mask = letters_to_mask(sistema, mecanismo_letters)
        escenarios.append((row, alcance_letters, mecanismo_letters, alcance_mask, mecanismo_mask))

    total = len(escenarios)
    print(f"Procesando {total} escenarios secuencialmente con aceleracion {modo}...")

    for idx, (row, alcance_letters, mecanismo_letters, alcance_mask, mecanismo_mask) in enumerate(escenarios, 1):
        t0 = time.perf_counter()
        print(f"\n[{idx:2d}/{total}] Row {row}: procesando...", end="", flush=True)
        results = _procesar_escenario(
            N, page, estado_inicial, condiciones_mask,
            alcance_mask, mecanismo_mask, use_gpu,
        )
        elapsed = time.perf_counter() - t0

        print(f"\n[{idx:2d}/{total}] Row {row}:")
        print(f"     alcance={alcance_letters} ({len(alcance_letters)}ch)")
        print(f"     mecanismo={mecanismo_letters} ({len(mecanismo_letters)}ch)")
        print(f"     tiempo_total={elapsed:.4f}s")

        for res in results:
            k = res['k']
            part_col, loss_col, time_col = K_COLUMNS[k]

            if 'error' in res:
                print(f"     k={k}: ERROR - {res['error']}")
                ws.cell(row=row, column=part_col, value=f"ERROR: {res['error']}")
                ws.cell(row=row, column=loss_col, value=None)
                ws.cell(row=row, column=time_col, value=None)
            else:
                ws.cell(row=row, column=part_col, value=res['particion'])
                ws.cell(row=row, column=loss_col, value=res['perdida'])
                ws.cell(row=row, column=time_col, value=round(res['tiempo'], 6))
                print(f"     k={k}: phi={res['perdida']:.6f}  t={res['tiempo']:.4f}s")

        if idx % 10 == 0:
            wb.save(excel_path)
            print(f"  (guardado parcial: {idx}/{total})")

    wb.save(excel_path)
    print(f"\nProcesados {total} escenarios. Guardado: {excel_path}")
